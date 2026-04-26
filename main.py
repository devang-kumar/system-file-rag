import os
import string
import datetime
import threading
import json
import numpy as np

from dotenv import load_dotenv
from fastapi import FastAPI
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_groq import ChatGroq

load_dotenv()

CHROMA_PATH = "./vector_db.json"
EMBED_MODEL = "voyage-3-lite"
GROQ_MODEL = "llama3-8b-8192"
MAX_FILE_SIZE_MB = 10
MAX_CONTENT_CHARS = 6000

SKIP_DIRS = {
    "Windows", "System32", "SysWOW64", "WinSxS", "$Recycle.Bin",
    "AppData", "ProgramData", "node_modules", "__pycache__",
    ".git", "venv", ".venv", "site-packages"
}

TEXT_EXTENSIONS = {
    ".txt", ".md", ".py", ".js", ".ts", ".jsx", ".tsx", ".html", ".css",
    ".json", ".xml", ".yaml", ".yml", ".csv", ".ini", ".cfg", ".toml",
    ".sh", ".bat", ".ps1", ".c", ".cpp", ".h", ".java", ".cs", ".go",
    ".rs", ".rb", ".php", ".sql", ".log", ".r"
}

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

vectorstore = {"docs": [], "embeddings": [], "metadatas": []}
llm = ChatGroq(model=GROQ_MODEL, api_key=os.getenv("GROQ_API_KEY"))

index_status = {"running": False, "done": False, "log": []}

try:
    import voyageai
    voyage_client = voyageai.Client(api_key=os.getenv("VOYAGE_API_KEY"))
except Exception:
    voyage_client = None


def get_drives():
    return [f"{l}:\\" for l in string.ascii_uppercase if os.path.exists(f"{l}:\\")]


def extract_content(path, ext):
    try:
        size = os.path.getsize(path)
        if size > MAX_FILE_SIZE_MB * 1024 * 1024:
            return ""

        if ext in TEXT_EXTENSIONS:
            import chardet
            with open(path, "rb") as f:
                raw = f.read(MAX_CONTENT_CHARS * 2)
            enc = chardet.detect(raw)["encoding"] or "utf-8"
            return raw.decode(enc, errors="ignore")[:MAX_CONTENT_CHARS]

        if ext == ".pdf":
            import fitz
            doc = fitz.open(path)
            text = ""
            for page in doc:
                text += page.get_text()
                if len(text) >= MAX_CONTENT_CHARS:
                    break
            return text[:MAX_CONTENT_CHARS]

        if ext in (".docx",):
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs)[:MAX_CONTENT_CHARS]

        if ext in (".xlsx",):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    text += " ".join(str(c) for c in row if c is not None) + "\n"
                    if len(text) >= MAX_CONTENT_CHARS:
                        return text[:MAX_CONTENT_CHARS]
            return text[:MAX_CONTENT_CHARS]

    except Exception:
        pass
    return ""


def build_text(path, name, ext, ftype, modified, size_kb, content):
    parent = os.path.basename(os.path.dirname(path))
    grandparent = os.path.basename(os.path.dirname(os.path.dirname(path)))
    text = (
        f"Name: {name}{ext}\n"
        f"Type: {ftype}\n"
        f"Extension: {ext}\n"
        f"Full path: {path}\n"
        f"Folder: {parent} inside {grandparent}\n"
        f"Last modified: {modified}\n"
        f"Size: {size_kb} KB\n"
    )
    if content:
        text += f"\nContent:\n{content}"
    return text


def cosine_similarity(a, b):
    return np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b))


def embed_texts(texts):
    if not voyage_client:
        return [[0.0] * 1024 for _ in texts]
    result = voyage_client.embed(texts, model=EMBED_MODEL)
    return result.embeddings


def run_indexing():
    global vectorstore
    index_status["running"] = True
    index_status["done"] = False
    index_status["log"] = ["Starting crawl..."]

    splitter = RecursiveCharacterTextSplitter(chunk_size=500, chunk_overlap=80)
    all_texts = []
    all_metas = []

    for drive in get_drives():
        index_status["log"].append(f"Crawling {drive}")
        for root, dirs, files in os.walk(drive, topdown=True):
            dirs[:] = [d for d in dirs if d not in SKIP_DIRS]

            entries = []
            for d in dirs:
                entries.append((os.path.join(root, d), d, "", "folder"))
            for f in files:
                name, ext = os.path.splitext(f)
                entries.append((os.path.join(root, f), name, ext.lower(), "file"))

            for path, name, ext, ftype in entries:
                try:
                    stat = os.stat(path)
                    size_kb = stat.st_size // 1024
                    modified = datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d")
                except Exception:
                    size_kb, modified = 0, ""

                content = extract_content(path, ext) if ftype == "file" else ""
                full_text = build_text(path, name, ext, ftype, modified, size_kb, content)
                chunks = splitter.split_text(full_text)

                for chunk in chunks:
                    all_texts.append(chunk)
                    all_metas.append({"path": path, "name": name, "ext": ext, "type": ftype, "modified": modified, "size_kb": size_kb})

                if len(all_texts) % 2000 == 0:
                    index_status["log"].append(f"  {len(all_texts)} chunks collected...")

    index_status["log"].append(f"Total chunks: {len(all_texts)}. Embedding...")

    batch_size = 128
    all_embeddings = []
    for i in range(0, len(all_texts), batch_size):
        batch = all_texts[i:i+batch_size]
        embs = embed_texts(batch)
        all_embeddings.extend(embs)
        index_status["log"].append(f"  Embedded {len(all_embeddings)}/{len(all_texts)}")

    vectorstore["docs"] = all_texts
    vectorstore["embeddings"] = all_embeddings
    vectorstore["metadatas"] = all_metas

    with open(CHROMA_PATH, "w") as f:
        json.dump({
            "docs": all_texts,
            "embeddings": [list(e) for e in all_embeddings],
            "metadatas": all_metas
        }, f)

    index_status["running"] = False
    index_status["done"] = True
    index_status["log"].append("Index complete. Ready to search.")


def load_existing_index():
    global vectorstore
    if os.path.exists(CHROMA_PATH):
        try:
            with open(CHROMA_PATH, "r") as f:
                data = json.load(f)
            vectorstore["docs"] = data["docs"]
            vectorstore["embeddings"] = [np.array(e) for e in data["embeddings"]]
            vectorstore["metadatas"] = data["metadatas"]
            index_status["done"] = True
            index_status["log"].append("Loaded existing index.")
        except Exception:
            pass


load_existing_index()


class ChatRequest(BaseModel):
    message: str


@app.get("/")
def serve_ui():
    return FileResponse("index.html")


@app.get("/status")
def get_status():
    return {
        "running": index_status["running"],
        "done": index_status["done"],
        "log": index_status["log"][-20:]
    }


@app.post("/index")
def start_index():
    if index_status["running"]:
        return {"message": "Already running."}
    t = threading.Thread(target=run_indexing, daemon=True)
    t.start()
    return {"message": "Indexing started."}


@app.post("/chat")
def chat(req: ChatRequest):
    if not index_status["done"]:
        return {"reply": "Index not built yet. Please build the index first from the sidebar.", "files": []}

    query_emb = embed_texts([req.message])[0]
    
    similarities = []
    for i, doc_emb in enumerate(vectorstore["embeddings"]):
        sim = cosine_similarity(np.array(query_emb), np.array(doc_emb))
        similarities.append((sim, i))
    
    similarities.sort(reverse=True)
    top_indices = [idx for _, idx in similarities[:8]]

    seen = {}
    for idx in top_indices:
        meta = vectorstore["metadatas"][idx]
        p = meta.get("path", "")
        if p not in seen:
            seen[p] = meta

    context_lines = []
    for meta in list(seen.values())[:6]:
        context_lines.append(
            f"- {meta.get('name','')}{meta.get('ext','')} | {meta.get('type','')} | {meta.get('modified','')} | {meta.get('size_kb','')} KB\n  Path: {meta.get('path','')}"
        )

    context = "\n".join(context_lines) if context_lines else "No matching files found."
    
    prompt = f"""You are a helpful file finder assistant. The user described what they're looking for.
Based on the search results below, provide a natural, conversational response.

User query: {req.message}

Relevant files found:
{context}

Respond naturally and helpfully."""

    reply = llm.invoke(prompt).content
    return {"reply": reply, "files": list(seen.values())[:6]}
